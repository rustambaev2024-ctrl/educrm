import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

try:
    from weasyprint import HTML
except Exception:  # pragma: no cover - optional runtime dependency behavior
    HTML = None


def _to_cell(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def export_excel(report_type: str, data: dict) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "Report Type"
    summary["B1"] = report_type
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)

    row = 3
    for key, value in data.items():
        if isinstance(value, list):
            continue
        summary.cell(row=row, column=1, value=key)
        summary.cell(row=row, column=2, value=_to_cell(value))
        row += 1

    for key, value in data.items():
        if not isinstance(value, list):
            continue
        sheet = workbook.create_sheet(title=key[:31] or "Data")
        if not value:
            sheet["A1"] = "No data"
            continue
        if isinstance(value[0], dict):
            headers = list(value[0].keys())
            for index, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=index, value=header).font = Font(bold=True)
            for row_index, item in enumerate(value, start=2):
                for col_index, header in enumerate(headers, start=1):
                    sheet.cell(row=row_index, column=col_index, value=_to_cell(item.get(header)))
        else:
            sheet["A1"] = key
            sheet["A1"].font = Font(bold=True)
            for row_index, item in enumerate(value, start=2):
                sheet.cell(row=row_index, column=1, value=_to_cell(item))

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _simple_pdf_bytes(text: str) -> bytes:
    safe_text = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    stream = f"BT /F1 11 Tf 40 760 Td ({safe_text}) Tj ET"
    objects = [
        "1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n",
        "2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n",
        (
            "3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
            "/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>\nendobj\n"
        ),
        "4 0 obj\n<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>\nendobj\n",
        f"5 0 obj\n<< /Length {len(stream)} >>\nstream\n{stream}\nendstream\nendobj\n",
    ]

    content = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(content))
        content.extend(obj.encode("latin-1"))

    xref_start = len(content)
    content.extend(f"xref\n0 {len(objects) + 1}\n".encode("latin-1"))
    content.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        content.extend(f"{offset:010d} 00000 n \n".encode("latin-1"))
    content.extend(
        (
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref_start}\n%%EOF"
        ).encode("latin-1")
    )
    return bytes(content)


def export_pdf(report_type: str, data: dict) -> bytes:
    body = json.dumps(data, indent=2, ensure_ascii=False)
    if HTML is not None:
        html = (
            "<html><body>"
            f"<h1>EduCRM report: {report_type}</h1>"
            f"<pre>{body}</pre>"
            "</body></html>"
        )
        try:
            return HTML(string=html).write_pdf()
        except Exception:
            pass
    return _simple_pdf_bytes(f"EduCRM report {report_type}. Data length: {len(body)}")

